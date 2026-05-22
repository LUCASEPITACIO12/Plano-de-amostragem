"""
excel_export.py
Gera o arquivo Excel do Plano de Amostragem no formato das concessões.
"""
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from calculos import (
    Sistema, LinhaPlano, gerar_plano, calc_anexo14, calc_psd,
    faixa_populacional, MESES,
)


# ── Estilos ───────────────────────────────────────────────────────────────────
AZ_ESC  = "1F3864"
AZ_MED  = "2E75B6"
AZ_CLA  = "DEEAF1"
VD_CLA  = "E2EFDA"
VD_ESC  = "375623"
AM_CLA  = "FFF2CC"
CZ_CLA  = "F2F2F2"
BRANCO  = "FFFFFF"

def fl(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def fn(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def bd(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def ac(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def al(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def hdr(c, t, bg=AZ_ESC, sz=10):
    c.value = t
    c.font  = fn(bold=True, color=BRANCO, size=sz)
    c.fill  = fl(bg)
    c.border = bd()
    c.alignment = ac()

def sec(c, t, bg=AZ_MED):
    c.value = t
    c.font  = fn(bold=True, color=BRANCO)
    c.fill  = fl(bg)
    c.border = bd()
    c.alignment = al()

def cel(c, v, bg=BRANCO, center=False, bold=False, color="000000", sz=10):
    c.value = v
    c.fill  = fl(bg)
    c.border = bd()
    c.alignment = ac() if center else al()
    c.font  = fn(bold=bold, color=color, size=sz)


def _apply_widths(ws, widths: list[int], hdrs: list[str]):
    """
    Aplica larguras com proteção contra desalinhamento.
    Se a lista de widths não casa com a de headers, levanta erro explícito.
    """
    if len(widths) != len(hdrs):
        raise ValueError(
            f"widths e headers desalinhados em '{ws.title}': "
            f"widths={len(widths)} hdrs={len(hdrs)}"
        )
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(j + 1)].width = w


# ── Aba PLANO RESUMIDO ────────────────────────────────────────────────────────

def aba_resumido(wb, sistemas: list[Sistema], ano: int):
    ws = wb.create_sheet("PLANO RESUMIDO")
    ws.freeze_panes = "A3"

    hdrs = [
        "ITEM",                       # 1
        "Município",                  # 2
        "Localidades",                # 3
        "Sistema",                    # 4
        "ETA / Unidade de Tratamento",# 5
        "Tipo de tratamento",         # 6
        "Tipo de captação",           # 7
        "Empresa Trat.",              # 8
        "Empresa Dist.",              # 9
        "Resp. Tratamento",           # 10
        "RT (Conselho/Nº)",           # 11
        "Nº TOTAL DE LIGAÇÕES ATIVAS",# 12
        "Pop. atendida",              # 13
        "Escopo responsabilidade",    # 14
        "Nº pontos rede (Anx.14)",    # 15
        "Faixa populacional",         # 16
        *[f"Total {m[:3]}" for m in MESES],  # 17-28
        "Total Anual (amostras/ano)", # 29
        "PSD freq.",                  # 30
        "PSD qtd/evento",             # 31
    ]

    ws.row_dimensions[1].height = 40
    for j, h in enumerate(hdrs):
        c = ws.cell(row=1, column=j + 1)
        hdr(c, h, bg=AZ_ESC if j >= 11 else AZ_MED)

    escopo_label = {
        "completo":  "Completo (captação + tratamento + distribuição)",
        "trat_dist": "Tratamento + Distribuição",
        "dist":      "Somente Distribuição",
    }

    for i, s in enumerate(sistemas):
        linhas = gerar_plano(s)
        psd    = calc_psd(s.tem_superficial, s.populacao)
        n_pts  = calc_anexo14(s.populacao)

        # Total por mês (excluindo operacionais)
        totais_mes = []
        for mes in range(1, 13):
            total = sum(
                l.quantidade_no_mes(mes) for l in linhas
                if not l.is_operacional
            )
            totais_mes.append(total)
        total_ano = sum(totais_mes)

        bg = AZ_CLA if i % 2 == 0 else BRANCO
        r  = i + 2
        ws.row_dimensions[r].height = 22

        row_data = [
            i + 1,
            s.municipio,
            s.localidades,
            s.nome,
            s.nome_eta,
            s.tratamento,
            s.manancial_efetivo,
            s.empresa_responsavel,
            s.empresa_distribuicao,
            s.responsavel_tratamento,
            f"{s.rt_nome} | {s.rt_conselho} {s.rt_registro}",
            s.n_ligacoes,
            s.populacao,
            escopo_label.get(s.escopo, s.escopo),
            n_pts,
            faixa_populacional(s.populacao),
            *totais_mes,
            total_ano,
            psd["freq"],
            psd["qtd"],
        ]

        for j, v in enumerate(row_data):
            c = ws.cell(row=r, column=j + 1)
            cel(c, v, bg=bg, center=(j >= 11))

    # Larguras alinhadas exatamente com hdrs (31 colunas)
    widths = [
        5,   # 1  ITEM
        14,  # 2  Município
        28,  # 3  Localidades
        24,  # 4  Sistema
        26,  # 5  ETA
        20,  # 6  Tipo de tratamento
        14,  # 7  Tipo de captação
        20,  # 8  Empresa Trat.
        20,  # 9  Empresa Dist.
        18,  # 10 Resp. Tratamento
        28,  # 11 RT
        12,  # 12 Nº ligações
        12,  # 13 Pop.
        36,  # 14 Escopo
        12,  # 15 Nº pontos rede
        22,  # 16 Faixa populacional
        *[7] * 12,  # 17-28 meses
        14,  # 29 Total Anual
        12,  # 30 PSD freq.
        10,  # 31 PSD qtd
    ]
    _apply_widths(ws, widths, hdrs)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"


# ── Aba PLANO ANUAL ───────────────────────────────────────────────────────────

def aba_anual(wb, sistemas: list[Sistema], ano: int):
    ws = wb.create_sheet("Plano - Anual")
    ws.freeze_panes = "A2"

    hdrs = [
        "Município", "Sistema", "Localidade", "Etapa",
        "Grupo de Parâmetros", "Parâmetro",
        "Ponto de coleta (tipo)", "Ponto de coleta (descrição)",
        "Frequência", "Qtd/evento",
        *[m[:3] for m in MESES],
        "Total Anual", "Base Legal", "OBS",
    ]

    ws.row_dimensions[1].height = 35
    for j, h in enumerate(hdrs):
        hdr(ws.cell(row=1, column=j + 1), h)

    row_num = 2
    for s in sistemas:
        linhas = gerar_plano(s)
        for li, l in enumerate(linhas):
            bg = AZ_CLA if li % 2 == 0 else BRANCO
            ws.row_dimensions[row_num].height = 20

            meses_qtd = [l.quantidade_no_mes(m) for m in range(1, 13)]

            row_data = [
                s.municipio, s.nome, s.localidades,
                l.etapa, l.grupo, l.parametro,
                l.ponto_tipo, l.ponto_desc,
                l.frequencia, l.quantidade,
                *meses_qtd,
                l.total_anual, l.base_legal, l.obs_ponto,
            ]

            for j, v in enumerate(row_data):
                c = ws.cell(row=row_num, column=j + 1)
                cel(c, v, bg=bg, center=(j >= 10))

            row_num += 1

    # 9 + 12 + 3 = 24 cabeçalhos? Vamos checar:
    # municipio, sistema, localidade, etapa, grupo, parametro,
    # ponto_tipo, ponto_desc, frequencia, qtd  = 10
    # + 12 meses = 22
    # + total, base, obs = 25
    widths = [
        14, 22, 22, 24, 28, 36, 20, 40, 18, 8,  # 10
        *[6] * 12,                                # 12 meses
        8, 18, 28,                                # 3
    ]
    _apply_widths(ws, widths, hdrs)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"


# ── Aba TAB RESUMO ────────────────────────────────────────────────────────────

def aba_tab_resumo(wb, sistemas: list[Sistema], ano: int):
    ws = wb.create_sheet("TAB Resumo")

    hdrs = ["Município", "Sistema", "Pop.", "Escopo",
            "Pontos rede", *[m[:3] for m in MESES], "Total Anual"]

    ws.row_dimensions[1].height = 35
    for j, h in enumerate(hdrs):
        hdr(ws.cell(row=1, column=j + 1), h)

    totais_gerais = [0] * 12
    total_geral_ano = 0

    for i, s in enumerate(sistemas):
        linhas = gerar_plano(s)
        n_pts  = calc_anexo14(s.populacao)
        escopo_label = {
            "completo":  "Completo",
            "trat_dist": "Trat.+Dist.",
            "dist":      "Só Dist.",
        }

        totais_mes = []
        for mes in range(1, 13):
            t = sum(l.quantidade_no_mes(mes) for l in linhas
                    if not l.is_operacional)
            totais_mes.append(t)
            totais_gerais[mes - 1] += t
        total_ano = sum(totais_mes)
        total_geral_ano += total_ano

        bg = VD_CLA if i % 2 == 0 else BRANCO
        r  = i + 2
        ws.row_dimensions[r].height = 22

        row_data = [
            s.municipio, s.nome, s.populacao,
            escopo_label.get(s.escopo, s.escopo), n_pts,
            *totais_mes, total_ano,
        ]
        for j, v in enumerate(row_data):
            c = ws.cell(row=r, column=j + 1)
            cel(c, v, bg=bg, center=(j >= 4))

    # Linha de total geral
    r_tot = len(sistemas) + 2
    ws.row_dimensions[r_tot].height = 28
    totals = ["TOTAL GERAL", "", "", "", "",
              *totais_gerais, total_geral_ano]
    for j, v in enumerate(totals):
        c = ws.cell(row=r_tot, column=j + 1)
        c.value = v
        c.fill  = fl(AZ_ESC)
        c.font  = fn(bold=True, color=BRANCO, size=11)
        c.border = bd()
        c.alignment = ac() if j >= 4 else al()

    widths = [14, 24, 10, 12, 10, *[8] * 12, 12]
    _apply_widths(ws, widths, hdrs)


# ── Aba MEMÓRIA DE CÁLCULO ────────────────────────────────────────────────────

def aba_memoria(wb, sistemas: list[Sistema]):
    """
    Explica COMO os números foram calculados. Útil para auditoria SESAU.
    """
    ws = wb.create_sheet("Memória de Cálculo")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    c = ws.cell(row=r, column=1)
    c.value = "Memória de Cálculo – Plano de Amostragem 888/2021"
    c.font = fn(bold=True, color=BRANCO, size=12)
    c.fill = fl(AZ_ESC); c.border = bd(); c.alignment = ac()
    ws.row_dimensions[r].height = 28
    r += 2

    for s in sistemas:
        # Cabeçalho do sistema
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(row=r, column=1)
        sec(c, f"{s.municipio} – {s.nome}")
        r += 1

        psd = calc_psd(s.tem_superficial, s.populacao)
        n_pts = calc_anexo14(s.populacao)

        linhas_mem = [
            ("População atendida", f"{s.populacao:,} hab.".replace(",", ".")),
            ("Faixa populacional (Anexo 14)", faixa_populacional(s.populacao)),
            ("Nº mínimo pontos rede", f"{n_pts} pontos (regra Anexo 14)"),
            ("Manancial efetivo", s.manancial_efetivo),
            ("Captações cadastradas",
             ", ".join(f"{c.nome} ({c.label_tipo})" for c in s.captacoes)),
            ("Desinfetante", s.desinfetante),
            ("Pré-oxidação", s.oxidante_preox),
            ("PSD – frequência", psd["freq"]),
            ("PSD – qtd por evento", str(psd["qtd"])),
            ("PSD – meses de coleta",
             ", ".join(MESES[m - 1] for m in psd["meses"])),
            ("Escopo de responsabilidade", {
                "completo":  "Completo",
                "trat_dist": "Tratamento + Distribuição",
                "dist":      "Somente Distribuição",
            }.get(s.escopo, s.escopo)),
            ("Horas/dia de operação", f"{s.horas_funcionamento:.0f} h"),
            ("Amostras 'a cada 2h' por dia",
             f"{int(s.horas_funcionamento / 2)} × por dia × dias do mês"),
        ]

        for label, valor in linhas_mem:
            cel(ws.cell(row=r, column=1), label, bg=CZ_CLA, bold=True)
            cel(ws.cell(row=r, column=2), valor, bg=BRANCO)
            ws.row_dimensions[r].height = 18
            r += 1
        r += 1  # linha em branco entre sistemas


# ── Aba REFERÊNCIA ANEXO 14 ───────────────────────────────────────────────────

def aba_anexo14(wb):
    ws = wb.create_sheet("Ref. Anexo 14")
    ws.row_dimensions[1].height = 40
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Anexo 14 – Nº Mínimo Mensal de Amostras Bacteriológicas na Rede (Portaria 888/2021)"
    c.font  = fn(bold=True, color=BRANCO, size=11)
    c.fill  = fl(AZ_ESC); c.border = bd(); c.alignment = ac()

    for j, h in enumerate(["Faixa Populacional", "Fórmula",
                            "Exemplo (pop.= valor médio)", "Nº mínimo"]):
        hdr(ws.cell(row=2, column=j + 1), h)

    faixas = [
        ("< 5.000 hab.",         "5",                             2_500,    5),
        ("5.000 a 10.000",       "10",                            7_000,   10),
        ("10.000 a 50.000",      "pop ÷ 1.000",                  30_000,  30),
        ("50.000 a 80.000",      "25 + pop ÷ 2.000",             65_000,  58),
        ("80.000 a 130.000",     "1 + pop ÷ 1.250",             105_000,  85),
        ("130.000 a 250.000",    "40 + pop ÷ 2.000",            190_000, 135),
        ("250.000 a 340.000",    "115 + pop ÷ 5.000",           295_000, 174),
        ("340.000 a 400.000",    "47 + pop ÷ 2.500",            370_000, 195),
        ("400.000 a 600.000",    "127 + pop ÷ 5.000",           500_000, 227),
        ("600.000 a 1.140.000",  "187 + pop ÷ 10.000",          870_000, 274),
        ("> 1.140.000 (máx.400)","min(400, 244 + pop ÷ 20.000)",1_500_000,319),
    ]
    for i, (faixa, formula, ex_pop, n) in enumerate(faixas):
        bg = AZ_CLA if i % 2 == 0 else BRANCO
        for j, v in enumerate([faixa, formula, f"{ex_pop:,} hab. → {n}", n]):
            c = ws.cell(row=i + 3, column=j + 1)
            cel(c, v, bg=bg, center=(j >= 2))

    for col, w in [("A", 30), ("B", 28), ("C", 30), ("D", 14)]:
        ws.column_dimensions[col].width = w


# ── Função principal ──────────────────────────────────────────────────────────

def gerar_excel(sistemas: list[Sistema], ano: int = 2026) -> bytes:
    """
    Gera o workbook completo e retorna os bytes para download.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove aba em branco padrão

    aba_resumido(wb, sistemas, ano)
    aba_anual(wb, sistemas, ano)
    aba_tab_resumo(wb, sistemas, ano)
    aba_memoria(wb, sistemas)
    aba_anexo14(wb)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
